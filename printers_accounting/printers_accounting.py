import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
from pathlib import Path
from datetime import datetime, timedelta
from uuid import uuid4
import json
import shutil
import os
import re


DATA_FILE = Path("consumables.json")
HISTORY_FILE = Path("history.json")
SETTINGS_FILE = Path("settings.json")
CATALOG_FILE = Path("catalog.json")
LISTS_FILE = Path("lists.json")
REPORTS_DIR = Path("reports")
BACKUPS_DIR = Path("backups")

AUTO_BACKUP_INTERVAL_MS = 5 * 60 * 1000
BACKUP_RETENTION_DAYS = 30
APP_FONT = "Segoe UI"

DEFAULT_WAREHOUSES = ["Этаж 8", "Этаж 11"]
DEFAULT_ITEM_TYPES = [
    "Картридж",
    "Драм-картридж",
    "Печь",
    "Принт-картридж",
    "Тонер",
    "Другое"
]
COLORS = ["Черный", "Голубой", "Пурпурный", "Желтый"]

THEME = {
    "bg": "#F6F7F9",
    "card": "#FFFFFF",
    "text": "#111827",
    "muted": "#6B7280",
    "border": "#E5E7EB",
    "input": "#F9FAFB",
    "header": "#F3F4F6",
    "row": "#FFFFFF",
    "row_alt": "#F9FAFB",
    "selected": "#2563EB",
    "selected_text": "#FFFFFF",
    "ok": "#DCFCE7",
    "low": "#FFF7D6",
    "empty": "#FEE2E2",
    "accent": "#2563EB",
    "accent_hover": "#1D4ED8",
    "button_gray": "#E5E7EB",
    "button_gray_hover": "#D1D5DB",
    "button_gray_text": "#111827",
    "danger": "#EF4444",
    "danger_hover": "#DC2626",
    "success": "#16A34A",
    "success_hover": "#15803D",
    "scrollbar": "#CBD5E1",
    "scrollbar_hover": "#94A3B8"
}

LOAD_WARNINGS = []


def read_json(path, default):
    if not path.exists():
        return default

    try:
        with open(path, "r", encoding="utf-8") as file:
            return json.load(file)
    except Exception as error:
        stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        bad_path = path.with_name(f"{path.stem}_поврежден_{stamp}{path.suffix}")

        try:
            shutil.copy2(path, bad_path)
            LOAD_WARNINGS.append(f"Файл {path.name} поврежден или не читается. Создана копия: {bad_path.name}")
        except Exception:
            LOAD_WARNINGS.append(f"Файл {path.name} поврежден или не читается. Копию создать не удалось.")

        return default


def write_json(path, data):
    tmp_path = path.with_name(f"{path.name}.tmp")

    with open(tmp_path, "w", encoding="utf-8") as file:
        json.dump(data, file, ensure_ascii=False, indent=4)
        file.flush()
        os.fsync(file.fileno())

    os.replace(tmp_path, path)


def to_int(value, default=0):
    try:
        return int(value)
    except Exception:
        return default


def now_text():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def natural_sort_key(text):
    parts = re.split(r"(\d+)", str(text).lower())
    return [int(part) if part.isdigit() else part for part in parts]


def sorted_values(values):
    result = []
    for value in values:
        value = str(value).strip()
        if value and value not in result:
            result.append(value)
    return sorted(result, key=natural_sort_key)


def add_unique(values, value):
    value = str(value).strip()
    if value and value not in values:
        values.append(value)
        return True
    return False


def get_status(quantity, min_quantity):
    quantity = to_int(quantity)
    min_quantity = to_int(min_quantity, 1)
    if quantity <= 0:
        return "Нет"
    if quantity <= min_quantity:
        return "Мало"
    return "Ок"


def smart_sort_value(value):
    value = str(value).strip().lower()

    if value == "":
        return 2, ""

    parts = re.split(r"(\d+)", value)
    result = []

    for part in parts:
        if part.isdigit():
            result.append(int(part))
        else:
            result.append(part)

    return 1, result


def safe_configure(widget, **kwargs):
    try:
        widget.configure(**kwargs)
    except Exception:
        pass


class AutoHideScrollbar(ctk.CTkScrollbar):
    def set(self, first, last):
        first = float(first)
        last = float(last)
        if first <= 0 and last >= 1:
            self.grid_remove()
        else:
            self.grid()
        super().set(first, last)


class App(ctk.CTk):
    def __init__(self):
        ctk.set_default_color_theme("blue")
        ctk.set_appearance_mode("light")
        super().__init__()

        self.items = []
        self.history = []
        self.settings = {}
        self.catalog = {}
        self.lists = {}
        self.selected_id = None
        self.warehouse_filter = "Все"
        self.stock_sort_column = None
        self.stock_sort_reverse = False
        self.history_sort_column = None
        self.history_sort_reverse = False

        self.cards = []
        self.text_labels = []
        self.muted_labels = []
        self.inputs = []
        self.buttons = []
        self.transparent_frames = []
        self.warehouse_buttons = {}
        self.stat_labels = {}

        self.stock_columns = [
            ("warehouse", "Этаж", 105),
            ("location", "Расположение", 170),
            ("type", "Расходник", 180),
            ("color", "Цвет", 120),
            ("article", "Артикул", 160),
            ("printer", "Принтер", 160),
            ("quantity", "Остаток", 100),
            ("min_quantity", "Мин.", 90),
            ("status", "Статус", 100)
        ]
        self.history_columns = [
            ("date", "Дата", 165),
            ("action", "Действие", 125),
            ("warehouse", "Этаж", 100),
            ("location", "Расположение", 180),
            ("type", "Расходник", 160),
            ("color", "Цвет", 105),
            ("article", "Артикул", 145),
            ("printer", "Принтер", 130),
            ("before", "Было", 80),
            ("after", "Стало", 80),
            ("change", "Изм.", 80),
            ("comment", "Комментарий", 360)
        ]

        self.load_all_data()
        self.build_window()
        self.build_ui()
        self.apply_theme()
        self.catalog_type_changed()
        self.clear_fields()
        self.show_load_warnings()
        self.after(AUTO_BACKUP_INTERVAL_MS, self.auto_backup)

    # ---------- данные ----------

    def theme(self):
        return THEME

    def save_items(self):
        write_json(DATA_FILE, self.items)

    def save_history(self):
        write_json(HISTORY_FILE, self.history)

    def save_settings(self):
        write_json(SETTINGS_FILE, self.settings)

    def save_catalog(self):
        write_json(CATALOG_FILE, self.catalog)

    def save_lists(self):
        write_json(LISTS_FILE, self.lists)

    def default_lists(self):
        return {
            "warehouses": DEFAULT_WAREHOUSES.copy(),
            "locations": {warehouse: [] for warehouse in DEFAULT_WAREHOUSES},
            "printers": [],
            "item_types": DEFAULT_ITEM_TYPES.copy(),
            "deleted": {
                "warehouses": [],
                "locations": [],
                "printers": [],
                "item_types": []
            }
        }

    def normalize_lists(self, data):
        result = self.default_lists()
        if not isinstance(data, dict):
            return result

        for warehouse in data.get("warehouses", []):
            add_unique(result["warehouses"], warehouse)

        locations = data.get("locations", {})
        if isinstance(locations, dict):
            for warehouse, values in locations.items():
                warehouse = str(warehouse).strip()
                if not warehouse:
                    continue
                add_unique(result["warehouses"], warehouse)
                result["locations"].setdefault(warehouse, [])
                for value in values if isinstance(values, list) else []:
                    add_unique(result["locations"][warehouse], value)
        elif isinstance(locations, list):
            for warehouse in result["warehouses"]:
                result["locations"].setdefault(warehouse, [])
            for value in locations:
                add_unique(result["locations"].setdefault("Этаж 8", []), value)

        for printer in data.get("printers", []):
            add_unique(result["printers"], printer)

        for item_type in data.get("item_types", []):
            add_unique(result["item_types"], item_type)

        deleted = data.get("deleted", {})
        if isinstance(deleted, dict):
            for key in ["warehouses", "locations", "printers", "item_types"]:
                for value in deleted.get(key, []):
                    add_unique(result["deleted"][key], value)

        self.remove_deleted_values_from_lists(result)
        result["warehouses"] = sorted_values(result["warehouses"])
        result["printers"] = sorted_values(result["printers"])
        result["item_types"] = sorted_values(result["item_types"])
        for warehouse in list(result["locations"].keys()):
            result["locations"][warehouse] = sorted_values(result["locations"].get(warehouse, []))
        return result

    def load_all_data(self):
        self.items = read_json(DATA_FILE, [])
        self.history = read_json(HISTORY_FILE, [])
        self.settings = read_json(SETTINGS_FILE, {"theme": "light"})
        self.catalog = read_json(CATALOG_FILE, {})
        self.lists = self.normalize_lists(read_json(LISTS_FILE, {}))

        if not isinstance(self.items, list):
            self.items = []
        if not isinstance(self.history, list):
            self.history = []
        if not isinstance(self.settings, dict):
            self.settings = {"theme": "light"}
        if not isinstance(self.catalog, dict):
            self.catalog = {}

        items_changed = False
        history_changed = False

        for item in self.items:
            if "id" not in item:
                item["id"] = str(uuid4())
                items_changed = True
            item.setdefault("warehouse", "Этаж 8")
            item.setdefault("location", item.get("printer_location", ""))
            item.setdefault("printer", item.get("model", ""))
            item.setdefault("type", "Картридж")
            item.setdefault("color", "-")
            item.setdefault("article", "")
            item["quantity"] = to_int(item.get("quantity", 0))
            item["min_quantity"] = to_int(item.get("min_quantity", 1), 1)

        for record in self.history:
            if "id" not in record:
                record["id"] = str(uuid4())
                history_changed = True
            record.setdefault("date", "")
            record.setdefault("action", "")
            record.setdefault("warehouse", "")
            record.setdefault("location", "")
            record.setdefault("printer", record.get("model", ""))
            record.setdefault("type", "")
            record.setdefault("color", "-")
            record.setdefault("article", "")
            record.setdefault("before", "")
            record.setdefault("after", "")
            record.setdefault("change", "")
            record.setdefault("comment", "")

        if items_changed:
            self.save_items()
        if history_changed:
            self.save_history()

        self.learn_lists_from_items()
        self.save_catalog()
        self.save_lists()

    def show_load_warnings(self):
        if LOAD_WARNINGS:
            messagebox.showwarning("Проблема чтения данных", "\n".join(LOAD_WARNINGS))

    def learn_lists_from_items(self):
        changed = False
        for item in self.items:
            warehouse = item.get("warehouse", "")
            location = item.get("location", "")
            printer = item.get("printer", "")
            item_type = item.get("type", "")

            if self.add_learned_value("warehouses", warehouse):
                changed = True
            if warehouse and self.add_learned_location(warehouse, location):
                changed = True
            if self.add_learned_value("printers", printer):
                changed = True
            if self.add_learned_value("item_types", item_type):
                changed = True

        self.remove_deleted_values_from_lists()
        if changed:
            self.sort_lists()

    def sort_lists(self):
        self.lists["warehouses"] = sorted_values(self.lists.get("warehouses", []))
        self.lists["printers"] = sorted_values(self.lists.get("printers", []))
        self.lists["item_types"] = sorted_values(self.lists.get("item_types", DEFAULT_ITEM_TYPES))
        for warehouse in list(self.lists.get("locations", {}).keys()):
            self.lists["locations"][warehouse] = sorted_values(self.lists["locations"].get(warehouse, []))

    def ensure_deleted_lists(self):
        self.lists.setdefault("deleted", {})
        for key in ["warehouses", "locations", "printers", "item_types"]:
            self.lists["deleted"].setdefault(key, [])

    def is_deleted_value(self, list_key, value):
        self.ensure_deleted_lists()
        value = str(value).strip()
        return value in self.lists["deleted"].get(list_key, [])

    def mark_deleted_value(self, list_key, value):
        self.ensure_deleted_lists()
        add_unique(self.lists["deleted"].setdefault(list_key, []), value)

    def unmark_deleted_value(self, list_key, value):
        self.ensure_deleted_lists()
        value = str(value).strip()
        while value in self.lists["deleted"].get(list_key, []):
            self.lists["deleted"][list_key].remove(value)

    def add_learned_value(self, list_key, value):
        if self.is_deleted_value(list_key, value):
            return False
        return add_unique(self.lists[list_key], value)

    def add_user_value(self, list_key, value):
        self.unmark_deleted_value(list_key, value)
        return add_unique(self.lists[list_key], value)

    def add_user_location(self, warehouse, location):
        self.unmark_deleted_value("warehouses", warehouse)
        self.unmark_deleted_value("locations", location)
        self.lists["locations"].setdefault(warehouse, [])
        return add_unique(self.lists["locations"][warehouse], location)

    def add_learned_location(self, warehouse, location):
        if self.is_deleted_value("warehouses", warehouse) or self.is_deleted_value("locations", location):
            return False
        self.lists["locations"].setdefault(warehouse, [])
        return add_unique(self.lists["locations"][warehouse], location)

    def remove_deleted_values_from_lists(self, lists_data=None):
        data = lists_data if lists_data is not None else self.lists
        deleted = data.get("deleted", {})

        for value in deleted.get("warehouses", []):
            while value in data.get("warehouses", []):
                data["warehouses"].remove(value)
            data.get("locations", {}).pop(value, None)

        for value in deleted.get("locations", []):
            for locations in data.get("locations", {}).values():
                while value in locations:
                    locations.remove(value)

        for value in deleted.get("printers", []):
            while value in data.get("printers", []):
                data["printers"].remove(value)

        for value in deleted.get("item_types", []):
            while value in data.get("item_types", []):
                data["item_types"].remove(value)

    # ---------- история ----------

    def log_history(self, action, item, before_qty, after_qty, change, comment=""):
        self.history.append({
            "id": str(uuid4()),
            "date": now_text(),
            "action": action,
            "warehouse": item.get("warehouse", ""),
            "location": item.get("location", ""),
            "printer": item.get("printer", ""),
            "type": item.get("type", ""),
            "color": item.get("color", "-"),
            "article": item.get("article", ""),
            "before": before_qty,
            "after": after_qty,
            "change": change,
            "comment": comment
        })
        self.save_history()

    def clear_history(self):
        answer = messagebox.askyesno(
            "Очистка истории",
            "Вы точно хотите полностью очистить историю изменений?\n\nПеред очисткой будет создан бэкап."
        )
        if not answer:
            return
        self.create_backup(show_message=False)
        self.history.clear()
        self.save_history()
        self.refresh_all()
        messagebox.showinfo("История", "История изменений очищена.")

    # ---------- справочник артикулов ----------

    def get_color_key(self, item_type, color):
        if item_type == "Принт-картридж":
            return color if color in COLORS else "Черный"
        return "-"

    def get_catalog_article(self, printer, item_type, color):
        printer = str(printer).strip()
        item_type = str(item_type).strip()
        color_key = self.get_color_key(item_type, color)
        if not printer or not item_type:
            return ""

        printer_data = self.catalog.get(printer, {})
        if isinstance(printer_data, dict):
            type_data = printer_data.get(item_type, {})
            if isinstance(type_data, dict) and type_data.get(color_key):
                return type_data[color_key]

        fallback = self.catalog.get("Общая модель", {})
        if isinstance(fallback, dict):
            fallback_type = fallback.get(item_type, {})
            if isinstance(fallback_type, dict):
                return fallback_type.get(color_key, "")
        return ""

    def remember_catalog_article(self, printer, item_type, color, article):
        printer = str(printer).strip()
        item_type = str(item_type).strip()
        article = str(article).strip()
        if not printer or not item_type or not article:
            return False
        color_key = self.get_color_key(item_type, color)
        self.catalog.setdefault(printer, {})
        self.catalog[printer].setdefault(item_type, {})
        if self.catalog[printer][item_type].get(color_key) == article:
            return False
        self.catalog[printer][item_type][color_key] = article
        self.save_catalog()
        return True

    def set_entry_text(self, entry, text):
        entry.delete(0, "end")
        entry.insert(0, text)

    def auto_fill_article(self):
        article = self.get_catalog_article(
            self.printer_box.get(),
            self.type_box.get(),
            self.color_box.get()
        )
        self.set_entry_text(self.article_entry, article)

    def on_warehouse_change(self, value=None):
        self.update_location_values()

    def on_printer_change(self, value=None):
        self.auto_fill_article()

    def on_type_change(self, value=None):
        if self.type_box.get() == "Принт-картридж":
            self.color_box.configure(state="readonly")
            if self.color_box.get() not in COLORS:
                self.color_box.set("Черный")
        else:
            self.color_box.set("-")
            self.color_box.configure(state="disabled")
        self.auto_fill_article()

    def on_color_change(self, value=None):
        self.auto_fill_article()

    def catalog_type_changed(self, value=None):
        if self.catalog_type_box.get() == "Принт-картридж":
            self.catalog_color_box.configure(state="readonly")
            if self.catalog_color_box.get() not in COLORS:
                self.catalog_color_box.set("Черный")
        else:
            self.catalog_color_box.set("-")
            self.catalog_color_box.configure(state="disabled")
        self.load_catalog_article_to_form()

    def load_catalog_article_to_form(self, value=None):
        article = self.get_catalog_article(
            self.catalog_printer_box.get(),
            self.catalog_type_box.get(),
            self.catalog_color_box.get()
        )
        self.set_entry_text(self.catalog_article_entry, article)

    def save_catalog_article(self):
        printer = self.catalog_printer_box.get().strip()
        item_type = self.catalog_type_box.get().strip()
        color = self.catalog_color_box.get().strip()
        article = self.catalog_article_entry.get().strip()
        if not printer or not item_type or not article:
            messagebox.showwarning("Ошибка", "Заполните принтер, тип расходника и артикул.")
            return
        self.add_user_value("printers", printer)
        self.add_user_value("item_types", item_type)
        self.sort_lists()
        self.save_lists()
        self.remember_catalog_article(printer, item_type, color, article)
        self.update_all_combo_values()
        self.update_catalog_preview()
        messagebox.showinfo("Справочник", "Артикул сохранен.")

    def delete_catalog_article(self):
        printer = self.catalog_printer_box.get().strip()
        item_type = self.catalog_type_box.get().strip()
        color_key = self.get_color_key(item_type, self.catalog_color_box.get().strip())
        article = self.catalog.get(printer, {}).get(item_type, {}).get(color_key, "")
        if not article:
            messagebox.showwarning("Ошибка", "Для выбранной связки артикул не найден.")
            return
        answer = messagebox.askyesno(
            "Удаление артикула",
            f"Удалить артикул «{article}»?\n\nПринтер: {printer}\nРасходник: {item_type}\nЦвет: {color_key}"
        )
        if not answer:
            return
        del self.catalog[printer][item_type][color_key]
        if not self.catalog[printer][item_type]:
            del self.catalog[printer][item_type]
        if not self.catalog[printer]:
            del self.catalog[printer]
        self.save_catalog()
        self.catalog_article_entry.delete(0, "end")
        self.auto_fill_article()
        self.update_catalog_preview()
        messagebox.showinfo("Справочник", "Артикул удален.")

    def update_catalog_preview(self):
        if not hasattr(self, "catalog_preview"):
            return
        if not self.catalog:
            self.catalog_preview.configure(text="Пока нет сохраненных артикулов.")
            return
        lines = []
        for printer in sorted_values(self.catalog.keys()):
            lines.append(f"Принтер: {printer}")
            printer_data = self.catalog.get(printer, {})
            if isinstance(printer_data, dict):
                for item_type in sorted_values(printer_data.keys()):
                    type_data = printer_data.get(item_type, {})
                    if not isinstance(type_data, dict):
                        continue
                    if item_type == "Принт-картридж":
                        for color in COLORS:
                            article = type_data.get(color, "")
                            if article:
                                lines.append(f"  {item_type} / {color}: {article}")
                    else:
                        article = type_data.get("-", "")
                        if article:
                            lines.append(f"  {item_type}: {article}")
            lines.append("")
        self.catalog_preview.configure(text="\n".join(lines).strip() or "Пока нет сохраненных артикулов.")

    # ---------- форма склада ----------

    def get_form_data(self):
        data = {
            "warehouse": self.warehouse_box.get().strip(),
            "location": self.location_box.get().strip(),
            "printer": self.printer_box.get().strip(),
            "type": self.type_box.get().strip(),
            "color": self.color_box.get().strip(),
            "article": self.article_entry.get().strip(),
            "quantity": self.quantity_entry.get().strip(),
            "min_quantity": self.min_quantity_entry.get().strip()
        }
        if not data["warehouse"] or not data["location"] or not data["printer"] or not data["type"]:
            messagebox.showwarning("Ошибка", "Заполните этаж, расположение, принтер и тип расходника.")
            return None
        if not data["quantity"].isdigit() or not data["min_quantity"].isdigit():
            messagebox.showwarning("Ошибка", "Количество и минимум должны быть числами.")
            return None
        if data["type"] != "Принт-картридж":
            data["color"] = "-"
        data["quantity"] = int(data["quantity"])
        data["min_quantity"] = int(data["min_quantity"])
        return data

    def remember_form_values(self, data):
        changed = False
        warehouse = data["warehouse"]

        if self.add_user_value("warehouses", warehouse):
            changed = True
        if self.add_user_location(warehouse, data["location"]):
            changed = True
        if self.add_user_value("printers", data["printer"]):
            changed = True
        if self.add_user_value("item_types", data["type"]):
            changed = True

        if changed:
            self.sort_lists()
            self.save_lists()
            self.update_all_combo_values()
        if self.remember_catalog_article(data["printer"], data["type"], data["color"], data["article"]):
            self.update_catalog_preview()

    def clear_fields(self, reset_selection=True):
        if reset_selection:
            self.selected_id = None
        self.warehouse_box.set("Этаж 8" if self.warehouse_filter == "Все" else self.warehouse_filter)
        self.update_location_values(clear_current=True)
        self.location_box.set("")
        self.printer_box.set("")
        self.type_box.set("Картридж")
        self.color_box.set("-")
        self.article_entry.delete(0, "end")
        self.quantity_entry.delete(0, "end")
        self.min_quantity_entry.delete(0, "end")
        self.on_type_change()
        if hasattr(self, "stock_tree"):
            self.stock_tree.selection_remove(self.stock_tree.selection())

    def fill_fields(self, item):
        self.clear_fields(reset_selection=False)
        self.warehouse_box.set(item.get("warehouse", "Этаж 8"))
        self.update_location_values(clear_current=False)
        self.location_box.set(item.get("location", ""))
        self.printer_box.set(item.get("printer", ""))
        self.type_box.set(item.get("type", "Картридж"))
        if item.get("type") == "Принт-картридж":
            self.color_box.configure(state="readonly")
            self.color_box.set(item.get("color", "Черный"))
        else:
            self.color_box.set("-")
            self.color_box.configure(state="disabled")
        self.article_entry.delete(0, "end")
        self.article_entry.insert(0, item.get("article", ""))
        self.quantity_entry.insert(0, item.get("quantity", 0))
        self.min_quantity_entry.insert(0, item.get("min_quantity", 1))

    def add_item(self):
        data = self.get_form_data()
        if data is None:
            return

        duplicate = self.find_duplicate_item(data)
        if duplicate:
            answer = messagebox.askyesno(
                "Позиция уже есть",
                "Такая позиция уже существует на складе.\n\n"
                f"Этаж: {data['warehouse']}\n"
                f"Расположение: {data['location']}\n"
                f"Принтер: {data['printer']}\n"
                f"Расходник: {data['type']}\n"
                f"Артикул: {data['article']}\n\n"
                f"Увеличить остаток существующей позиции на {data['quantity']}?"
            )
            if not answer:
                return

            before_qty = to_int(duplicate.get("quantity", 0))
            duplicate["quantity"] = before_qty + data["quantity"]
            duplicate["min_quantity"] = data["min_quantity"]
            self.selected_id = duplicate.get("id")
            self.save_items()
            self.remember_form_values(data)
            self.log_history("Приход", duplicate, before_qty, duplicate["quantity"], data["quantity"], "Количество добавлено к существующей позиции")
            self.clear_fields()
            self.refresh_all()
            return

        data["id"] = str(uuid4())
        self.items.append(data)
        self.selected_id = None
        self.save_items()
        self.remember_form_values(data)
        self.log_history("Добавление", data, 0, data["quantity"], data["quantity"], "Добавлен новый расходник")
        self.clear_fields()
        self.refresh_all()

    def update_item(self):
        if self.selected_id is None:
            messagebox.showwarning("Ошибка", "Выберите строку.")
            return
        index = self.find_item_index(self.selected_id)
        if index is None:
            messagebox.showwarning("Ошибка", "Выбранная запись не найдена.")
            return
        data = self.get_form_data()
        if data is None:
            return

        duplicate = self.find_duplicate_item(data, exclude_id=self.selected_id)
        if duplicate:
            messagebox.showwarning(
                "Дубликат",
                "Другая такая же позиция уже есть в таблице.\n\n"
                "Измените этаж, расположение, принтер, тип, цвет или артикул, либо удалите лишнюю строку."
            )
            return

        before_qty = to_int(self.items[index].get("quantity", 0))
        after_qty = data["quantity"]
        data["id"] = self.selected_id
        self.items[index] = data
        self.save_items()
        self.remember_form_values(data)
        self.log_history("Изменение", data, before_qty, after_qty, after_qty - before_qty, "Изменены данные расходника")
        self.clear_fields()
        self.refresh_all()

    def delete_item(self):
        if self.selected_id is None:
            messagebox.showwarning("Ошибка", "Выберите строку.")
            return
        index = self.find_item_index(self.selected_id)
        if index is None:
            messagebox.showwarning("Ошибка", "Выбранная запись не найдена.")
            return
        item = self.items[index]
        quantity = to_int(item.get("quantity", 0))
        if not messagebox.askyesno("Удаление", "Удалить выбранный расходник?"):
            return
        self.create_backup(show_message=False)
        self.log_history("Удаление", item, quantity, 0, -quantity, "Расходник удален")
        self.items.pop(index)
        self.selected_id = None
        self.save_items()
        self.clear_fields()
        self.refresh_all()

    def change_quantity(self, sign):
        if self.selected_id is None:
            messagebox.showwarning("Ошибка", "Выберите строку.")
            return

        item = self.find_item_by_id(self.selected_id)

        if item is None:
            messagebox.showwarning("Ошибка", "Выбранная запись не найдена.")
            return

        before = to_int(item.get("quantity", 0))
        after = max(before + sign, 0)
        change = after - before

        if change == 0:
            messagebox.showinfo("Списание", "Остаток уже равен нулю.")
            return

        item["quantity"] = after
        self.save_items()

        action = "Приход" if sign > 0 else "Списание"
        self.log_history(action, item, before, after, change, f"{action.lower()} расходника на 1")
        self.selected_id = item.get("id")
        self.refresh_all()

    def clear_form(self):
        self.clear_fields()
        self.refresh_all()

    def find_item_by_id(self, item_id):
        for item in self.items:
            if item.get("id") == item_id:
                return item
        return None

    def find_item_index(self, item_id):
        for index, item in enumerate(self.items):
            if item.get("id") == item_id:
                return index
        return None

    def same_stock_position(self, item, data):
        keys = ["warehouse", "location", "printer", "type", "color", "article"]
        return all(str(item.get(key, "")).strip() == str(data.get(key, "")).strip() for key in keys)

    def find_duplicate_item(self, data, exclude_id=None):
        for item in self.items:
            if exclude_id is not None and item.get("id") == exclude_id:
                continue
            if self.same_stock_position(item, data):
                return item
        return None

    # ---------- фильтры ----------

    def get_stock_records(self):
        records = []
        search = self.stock_search_entry.get().strip().lower()
        type_filter = self.stock_type_filter.get()
        color_filter = self.stock_color_filter.get()
        status_filter = self.stock_status_filter.get()

        for item in self.items:
            if self.warehouse_filter != "Все" and item.get("warehouse") != self.warehouse_filter:
                continue
            if type_filter != "Все" and item.get("type") != type_filter:
                continue
            if color_filter != "Все" and item.get("color", "-") != color_filter:
                continue
            status = get_status(item.get("quantity", 0), item.get("min_quantity", 1))
            if status_filter != "Все" and status != status_filter:
                continue
            text = " ".join(str(item.get(key, "")) for key in [
                "warehouse", "location", "printer", "type", "color", "article"
            ]).lower()
            if search and search not in text:
                continue
            records.append(item)

        if self.stock_sort_column:
            records.sort(key=lambda item: smart_sort_value(self.get_item_value(item, self.stock_sort_column)), reverse=self.stock_sort_reverse)
        return records

    def get_history_records(self):
        records = []
        search = self.history_search_entry.get().strip().lower()
        action_filter = self.history_action_filter.get()
        for record in self.history:
            if action_filter != "Все" and record.get("action") != action_filter:
                continue
            text = " ".join(str(record.get(key, "")) for key in [
                "date", "action", "warehouse", "location", "printer", "type", "color", "article", "comment"
            ]).lower()
            if search and search not in text:
                continue
            records.append(record)
        if self.history_sort_column:
            records.sort(key=lambda record: smart_sort_value(record.get(self.history_sort_column, "")), reverse=self.history_sort_reverse)
        return records

    def get_item_value(self, item, column):
        if column == "status":
            return get_status(item.get("quantity", 0), item.get("min_quantity", 1))
        return item.get(column, "")

    def set_warehouse_filter(self, value):
        self.warehouse_filter = value
        if value != "Все":
            self.warehouse_box.set(value)
            self.update_location_values(clear_current=True)
        self.update_warehouse_buttons()
        self.refresh_all()

    def sort_stock(self, column):
        if self.stock_sort_column == column:
            self.stock_sort_reverse = not self.stock_sort_reverse
        else:
            self.stock_sort_column = column
            self.stock_sort_reverse = False
        self.refresh_stock_table()

    def sort_history(self, column):
        if self.history_sort_column == column:
            self.history_sort_reverse = not self.history_sort_reverse
        else:
            self.history_sort_column = column
            self.history_sort_reverse = False
        self.refresh_history_table()

    # ---------- бэкап и экспорт ----------

    def cleanup_old_backups(self):
        if not BACKUPS_DIR.exists():
            return
        cutoff = datetime.now() - timedelta(days=BACKUP_RETENTION_DAYS)
        for file in BACKUPS_DIR.glob("*.json"):
            try:
                if datetime.fromtimestamp(file.stat().st_mtime) < cutoff:
                    file.unlink()
            except Exception:
                pass

    def create_backup(self, show_message=True):
        BACKUPS_DIR.mkdir(exist_ok=True)
        self.cleanup_old_backups()
        stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        copied = 0
        for file in [DATA_FILE, HISTORY_FILE, SETTINGS_FILE, CATALOG_FILE, LISTS_FILE]:
            if file.exists():
                shutil.copy2(file, BACKUPS_DIR / f"{file.stem}_{stamp}{file.suffix}")
                copied += 1
        self.cleanup_old_backups()
        if show_message:
            messagebox.showinfo(
                "Резервная копия",
                f"Создано файлов: {copied}\nСтарые бэкапы старше {BACKUP_RETENTION_DAYS} дней удалены.\nПапка: {BACKUPS_DIR.resolve()}"
            )
        self.update_reports_info()

    def auto_backup(self):
        self.create_backup(show_message=False)
        self.after(AUTO_BACKUP_INTERVAL_MS, self.auto_backup)

    def close_app(self):
        try:
            self.create_backup(show_message=False)
        finally:
            self.destroy()

    def export_items_to_excel(self, records, filename_prefix, sheet_name):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except Exception:
            messagebox.showerror("Ошибка", "Для экспорта установите библиотеку:\n\npip install openpyxl")
            return

        REPORTS_DIR.mkdir(exist_ok=True)
        path = REPORTS_DIR / f"{filename_prefix}_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

        headers = ["Этаж", "Расположение", "Расходник", "Цвет", "Артикул", "Принтер", "Остаток", "Мин. остаток", "Статус"]
        sheet.append(headers)

        for item in records:
            sheet.append([
                item.get("warehouse", ""),
                item.get("location", ""),
                item.get("type", ""),
                item.get("color", "-"),
                item.get("article", ""),
                item.get("printer", ""),
                item.get("quantity", 0),
                item.get("min_quantity", 1),
                get_status(item.get("quantity", 0), item.get("min_quantity", 1))
            ])

        header_fill = PatternFill("solid", fgColor="E5E7EB")
        border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB")
        )

        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", horizontal="center")

        for column in sheet.columns:
            letter = column[0].column_letter
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            sheet.column_dimensions[letter].width = min(max_len + 3, 35)

        sheet.freeze_panes = "A2"
        workbook.save(path)
        messagebox.showinfo("Экспорт", f"Файл создан:\n{path.resolve()}")
        self.update_reports_info()

    def export_stock(self):
        self.export_items_to_excel(self.get_stock_records(), "Склад_расходников", "Склад")

    def export_shortage(self):
        records = []
        for item in self.get_stock_records():
            status = get_status(item.get("quantity", 0), item.get("min_quantity", 1))
            if status in ["Мало", "Нет"]:
                records.append(item)

        if not records:
            messagebox.showinfo("Дефицит", "Позиций со статусом «Мало» или «Нет» не найдено.")
            return

        self.export_items_to_excel(records, "Дефицит_расходников", "Дефицит")

    def restore_from_backup(self):
        if not BACKUPS_DIR.exists():
            messagebox.showwarning("Бэкап", "Папка backups пока не создана.")
            return

        selected = filedialog.askopenfilename(
            title="Выберите любой JSON-файл из нужного бэкапа",
            initialdir=str(BACKUPS_DIR.resolve()),
            filetypes=[("JSON backup", "*.json")]
        )

        if not selected:
            return

        selected_path = Path(selected)
        stamp = None
        source_files = [DATA_FILE, HISTORY_FILE, SETTINGS_FILE, CATALOG_FILE, LISTS_FILE]

        for original in source_files:
            prefix = f"{original.stem}_"
            if selected_path.stem.startswith(prefix):
                stamp = selected_path.stem[len(prefix):]
                break

        if not stamp:
            messagebox.showwarning("Бэкап", "Не удалось определить дату бэкапа по имени файла.")
            return

        available = []
        for original in source_files:
            backup_file = BACKUPS_DIR / f"{original.stem}_{stamp}{original.suffix}"
            if backup_file.exists():
                available.append((backup_file, original))

        if not available:
            messagebox.showwarning("Бэкап", "Файлы выбранного бэкапа не найдены.")
            return

        names = "\n".join(original.name for _, original in available)
        answer = messagebox.askyesno(
            "Восстановление из бэкапа",
            f"Восстановить данные из бэкапа {stamp}?\n\nБудут восстановлены файлы:\n{names}\n\nПеред восстановлением будет создан текущий бэкап."
        )

        if not answer:
            return

        self.create_backup(show_message=False)

        try:
            for backup_file, original in available:
                shutil.copy2(backup_file, original)
        except Exception as error:
            messagebox.showerror("Бэкап", f"Не удалось восстановить бэкап:\n{error}")
            return

        self.load_all_data()
        self.update_all_combo_values()
        self.clear_fields()
        self.refresh_all()
        self.update_catalog_preview()
        messagebox.showinfo("Бэкап", "Данные восстановлены из бэкапа.")

    def open_folder(self, path):
        path.mkdir(exist_ok=True)
        try:
            os.startfile(path.resolve())
        except Exception:
            messagebox.showinfo("Папка", str(path.resolve()))

    def open_data_folder(self):
        try:
            os.startfile(Path.cwd())
        except Exception:
            messagebox.showinfo("Папка с данными", str(Path.cwd()))

    # ---------- UI ----------

    def build_window(self):
        self.title("Учет расходников")
        self.geometry("1500x900")
        self.minsize(1360, 850)
        self.protocol("WM_DELETE_WINDOW", self.close_app)

    def build_ui(self):
        self.build_header()
        self.tab_view = ctk.CTkTabview(self, corner_radius=18)
        self.tab_view.pack(fill="both", expand=True, padx=22, pady=(6, 18))
        self.stock_tab = self.tab_view.add("Склад")
        self.history_tab = self.tab_view.add("История")
        self.reports_tab = self.tab_view.add("Отчеты")
        self.settings_tab = self.tab_view.add("Настройки")
        self.build_stock_tab()
        self.build_history_tab()
        self.build_reports_tab()
        self.build_settings_tab()

    def create_card(self, parent, height=None, fill="x", expand=False, padx=22, pady=8, border=True):
        p = self.theme()
        card = ctk.CTkFrame(
            parent,
            fg_color=p["card"],
            corner_radius=22,
            border_width=1 if border else 0,
            border_color=p["border"]
        )
        if height:
            card.configure(height=height)
            card.pack_propagate(False)
        card.pack(fill=fill, expand=expand, padx=padx, pady=pady)
        self.cards.append(card)
        return card

    def create_label(self, parent, text, row, column, kind="muted"):
        p = self.theme()
        label = ctk.CTkLabel(parent, text=text, font=(APP_FONT, 13), text_color=p["muted"] if kind == "muted" else p["text"])
        label.grid(row=row, column=column, padx=10, pady=(8, 3), sticky="w")
        (self.muted_labels if kind == "muted" else self.text_labels).append(label)
        return label

    def create_entry(self, parent, row, column, width=180):
        p = self.theme()
        entry = ctk.CTkEntry(parent, width=width, height=38, corner_radius=12, font=(APP_FONT, 13), fg_color=p["input"], border_color=p["border"], text_color=p["text"])
        entry.grid(row=row, column=column, padx=10, pady=(0, 10), sticky="w")
        self.inputs.append(entry)
        return entry

    def create_combo(self, parent, values, row, column, width=180, command=None, state="normal"):
        p = self.theme()
        combo = ctk.CTkComboBox(parent, values=values, width=width, height=38, corner_radius=12, font=(APP_FONT, 13), command=command, state=state, fg_color=p["input"], border_color=p["border"], text_color=p["text"])
        combo.grid(row=row, column=column, padx=10, pady=(0, 10), sticky="w")
        self.inputs.append(combo)
        return combo

    def create_button(self, parent, text, command, kind="accent", side="left", width=145):
        p = self.theme()
        colors = {
            "danger": (p["danger"], p["danger_hover"], "#FFFFFF"),
            "success": (p["success"], p["success_hover"], "#FFFFFF"),
            "gray": (p["button_gray"], p["button_gray_hover"], p["button_gray_text"]),
            "accent": (p["accent"], p["accent_hover"], "#FFFFFF")
        }
        fg, hover, text_color = colors.get(kind, colors["accent"])
        button = ctk.CTkButton(parent, text=text, command=command, width=width, height=40, corner_radius=14, font=(APP_FONT, 13, "bold"), fg_color=fg, hover_color=hover, text_color=text_color)
        button.pack(side=side, padx=6, pady=4)
        self.buttons.append((button, kind))
        return button

    def build_header(self):
        self.top_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.top_frame.pack(fill="x", padx=28, pady=(20, 4))
        self.transparent_frames.append(self.top_frame)

        title = ctk.CTkLabel(self.top_frame, text="Учет расходников", font=(APP_FONT, 30, "bold"))
        title.pack(side="left")
        self.text_labels.append(title)

        subtitle = ctk.CTkLabel(self.top_frame, text="Склад, история движения, отчеты и справочник артикулов", font=(APP_FONT, 14))
        subtitle.pack(side="left", padx=18, pady=(8, 0))
        self.muted_labels.append(subtitle)

    def build_stock_tab(self):
        self.stats_frame = ctk.CTkFrame(self.stock_tab, fg_color="transparent")
        self.stats_frame.pack(fill="x", padx=8, pady=(8, 0))
        self.transparent_frames.append(self.stats_frame)
        for key, title in [("total", "Позиций"), ("low", "Мало"), ("empty", "Нет в наличии"), ("floor", "Текущий этаж")]:
            self.create_stat_card(key, title)

        form_card = self.create_card(self.stock_tab, height=166)
        self.create_label(form_card, "Этаж", 0, 0)
        self.warehouse_box = self.create_combo(form_card, self.warehouse_values(), 1, 0, width=150, command=self.on_warehouse_change)
        self.warehouse_box.set("Этаж 8")

        self.create_label(form_card, "Расположение", 0, 1)
        self.location_box = self.create_combo(form_card, [], 1, 1, width=185)

        self.create_label(form_card, "Принтер", 0, 2)
        self.printer_box = self.create_combo(form_card, self.printer_values(), 1, 2, width=185, command=self.on_printer_change)

        self.create_label(form_card, "Тип", 0, 3)
        self.type_box = self.create_combo(form_card, self.type_values(), 1, 3, width=185, command=self.on_type_change)
        self.type_box.set("Картридж")

        self.create_label(form_card, "Цвет", 0, 4)
        self.color_box = self.create_combo(form_card, COLORS, 1, 4, width=150, command=self.on_color_change, state="readonly")
        self.color_box.set("-")
        self.color_box.configure(state="disabled")

        self.create_label(form_card, "Артикул", 2, 0)
        self.article_entry = self.create_entry(form_card, 3, 0, width=185)
        self.create_label(form_card, "Количество", 2, 1)
        self.quantity_entry = self.create_entry(form_card, 3, 1, width=185)
        self.create_label(form_card, "Мин. остаток", 2, 2)
        self.min_quantity_entry = self.create_entry(form_card, 3, 2, width=185)

        hint = ctk.CTkLabel(form_card, text="Новые этажи, расположения, принтеры и типы запоминаются после добавления или сохранения позиции.", font=(APP_FONT, 13), justify="left")
        hint.grid(row=3, column=3, columnspan=3, padx=10, pady=(0, 10), sticky="w")
        self.muted_labels.append(hint)

        filter_card = self.create_card(self.stock_tab, height=94)
        self.create_label(filter_card, "Поиск", 0, 0)
        self.stock_search_entry = self.create_entry(filter_card, 1, 0, width=260)
        self.stock_search_entry.bind("<KeyRelease>", lambda event: self.refresh_all())

        self.create_label(filter_card, "Тип", 0, 1)
        self.stock_type_filter = self.create_combo(filter_card, ["Все"] + self.type_values(), 1, 1, width=170, command=lambda value: self.refresh_all(), state="readonly")
        self.stock_type_filter.set("Все")

        self.create_label(filter_card, "Цвет", 0, 2)
        self.stock_color_filter = self.create_combo(filter_card, ["Все", "-"] + COLORS, 1, 2, width=150, command=lambda value: self.refresh_all(), state="readonly")
        self.stock_color_filter.set("Все")

        self.create_label(filter_card, "Статус", 0, 3)
        self.stock_status_filter = self.create_combo(filter_card, ["Все", "Ок", "Мало", "Нет"], 1, 3, width=150, command=lambda value: self.refresh_all(), state="readonly")
        self.stock_status_filter.set("Все")

        self.warehouse_switch_frame = ctk.CTkFrame(filter_card, fg_color="transparent")
        self.warehouse_switch_frame.grid(row=1, column=4, padx=18, pady=(0, 10), sticky="w")
        self.transparent_frames.append(self.warehouse_switch_frame)
        self.rebuild_warehouse_buttons()

        self.button_frame = ctk.CTkFrame(self.stock_tab, fg_color="transparent")
        self.button_frame.pack(fill="x", padx=14, pady=(0, 4))
        self.transparent_frames.append(self.button_frame)
        self.create_button(self.button_frame, "Добавить", self.add_item)
        self.create_button(self.button_frame, "Сохранить", self.update_item)
        self.create_button(self.button_frame, "Приход", lambda: self.change_quantity(1), kind="gray")
        self.create_button(self.button_frame, "Списание", lambda: self.change_quantity(-1), kind="gray")
        self.create_button(self.button_frame, "Очистить", self.clear_form, kind="gray")
        self.create_button(self.button_frame, "Удалить", self.delete_item, kind="danger", side="right")

        table_card = self.create_card(self.stock_tab, fill="both", expand=True, border=False)
        self.stock_tree_frame = ctk.CTkFrame(table_card, fg_color="transparent")
        self.stock_tree_frame.pack(fill="both", expand=True, padx=14, pady=14)
        self.transparent_frames.append(self.stock_tree_frame)
        self.stock_tree = self.create_tree(self.stock_tree_frame, self.stock_columns, self.sort_stock)
        self.stock_tree.bind("<<TreeviewSelect>>", self.stock_selection_changed)
        self.update_location_values(clear_current=True)

    def create_stat_card(self, key, title):
        p = self.theme()
        card = ctk.CTkFrame(self.stats_frame, fg_color=p["card"], corner_radius=18, border_width=1, border_color=p["border"], width=190, height=72)
        card.pack(side="left", padx=8, pady=6)
        card.pack_propagate(False)
        self.cards.append(card)
        caption = ctk.CTkLabel(card, text=title, font=(APP_FONT, 12), text_color=p["muted"])
        caption.pack(anchor="w", padx=16, pady=(10, 0))
        self.muted_labels.append(caption)
        value = ctk.CTkLabel(card, text="0", font=(APP_FONT, 24, "bold"), text_color=p["text"])
        value.pack(anchor="w", padx=16, pady=(0, 8))
        self.text_labels.append(value)
        self.stat_labels[key] = value

    def create_tree(self, parent, columns, sort_command):
        tree = ttk.Treeview(parent, columns=[column[0] for column in columns], show="headings", selectmode="browse", style="Clean.Treeview")
        y_scroll = AutoHideScrollbar(parent, orientation="vertical", command=tree.yview, width=12)
        x_scroll = AutoHideScrollbar(parent, orientation="horizontal", command=tree.xview, height=12)
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns", padx=(6, 0))
        x_scroll.grid(row=1, column=0, sticky="ew", pady=(6, 0))
        parent.grid_rowconfigure(0, weight=1)
        parent.grid_columnconfigure(0, weight=1)
        for column, title, width in columns:
            tree.heading(column, text=title, anchor="center", command=lambda c=column: sort_command(c))
            tree.column(column, width=width, minwidth=70, stretch=True, anchor="center")
        if columns is self.stock_columns:
            self.stock_y_scroll, self.stock_x_scroll = y_scroll, x_scroll
        else:
            self.history_y_scroll, self.history_x_scroll = y_scroll, x_scroll
        return tree

    def get_stock_row_tag(self, item, index):
        status = get_status(item.get("quantity", 0), item.get("min_quantity", 1))

        if status == "Нет":
            return "empty"
        if status == "Мало":
            return "low"
        if status == "Ок":
            return "ok"

        return "even" if index % 2 == 0 else "odd"

    def refresh_stock_selection_style(self):
        if not hasattr(self, "stock_tree"):
            return

        for index, item_id in enumerate(self.stock_tree.get_children()):
            if item_id == self.selected_id:
                self.stock_tree.item(item_id, tags=("selected_row",))
            else:
                item = self.find_item_by_id(item_id)
                if item:
                    self.stock_tree.item(item_id, tags=(self.get_stock_row_tag(item, index),))

    def stock_selection_changed(self, event=None):
        selected = self.stock_tree.selection()
        if not selected:
            return

        self.selected_id = selected[0]
        self.stock_tree.focus(self.selected_id)
        self.stock_tree.see(self.selected_id)
        self.refresh_stock_selection_style()

        item = self.find_item_by_id(self.selected_id)

        if item:
            self.fill_fields(item)

    def build_history_tab(self):
        card = self.create_card(self.history_tab, height=94)
        self.create_label(card, "Поиск по истории", 0, 0)
        self.history_search_entry = self.create_entry(card, 1, 0, width=320)
        self.history_search_entry.bind("<KeyRelease>", lambda event: self.refresh_history_table())
        self.create_label(card, "Действие", 0, 1)
        self.history_action_filter = self.create_combo(card, ["Все", "Добавление", "Изменение", "Приход", "Списание", "Удаление"], 1, 1, width=190, command=lambda value: self.refresh_history_table(), state="readonly")
        self.history_action_filter.set("Все")
        table_card = self.create_card(self.history_tab, fill="both", expand=True, border=False)
        self.history_tree_frame = ctk.CTkFrame(table_card, fg_color="transparent")
        self.history_tree_frame.pack(fill="both", expand=True, padx=14, pady=14)
        self.transparent_frames.append(self.history_tree_frame)
        self.history_tree = self.create_tree(self.history_tree_frame, self.history_columns, self.sort_history)

    def build_reports_tab(self):
        card = self.create_card(self.reports_tab, height=250)
        title = ctk.CTkLabel(card, text="Отчеты и резервные копии", font=(APP_FONT, 22, "bold"))
        title.pack(anchor="w", padx=22, pady=(20, 6))
        self.text_labels.append(title)
        text = ctk.CTkLabel(card, text="Здесь можно выгрузить склад в Excel и создать резервную копию данных.", font=(APP_FONT, 14))
        text.pack(anchor="w", padx=22, pady=(0, 16))
        self.muted_labels.append(text)
        self.reports_buttons_frame = ctk.CTkFrame(card, fg_color="transparent")
        self.reports_buttons_frame.pack(anchor="w", padx=16, pady=4)
        self.transparent_frames.append(self.reports_buttons_frame)
        self.create_button(self.reports_buttons_frame, "Экспорт склада", self.export_stock, kind="success", width=170)
        self.create_button(self.reports_buttons_frame, "Экспорт дефицита", self.export_shortage, kind="success", width=180)
        self.create_button(self.reports_buttons_frame, "Создать бэкап", self.create_backup, kind="gray", width=160)
        self.create_button(self.reports_buttons_frame, "Восстановить бэкап", self.restore_from_backup, kind="gray", width=190)
        self.create_button(self.reports_buttons_frame, "Папка отчетов", lambda: self.open_folder(REPORTS_DIR), kind="gray", width=160)
        self.create_button(self.reports_buttons_frame, "Папка бэкапов", lambda: self.open_folder(BACKUPS_DIR), kind="gray", width=160)
        self.reports_info_label = ctk.CTkLabel(card, text="", font=(APP_FONT, 13))
        self.reports_info_label.pack(anchor="w", padx=22, pady=(12, 0))
        self.muted_labels.append(self.reports_info_label)

    def build_settings_tab(self):
        settings_card = self.create_card(self.settings_tab, height=170)
        title = ctk.CTkLabel(settings_card, text="Настройки", font=(APP_FONT, 22, "bold"))
        title.pack(anchor="w", padx=22, pady=(18, 6))
        self.text_labels.append(title)
        text = ctk.CTkLabel(
            settings_card,
            text="Автобэкап создается каждые 5 минут и при закрытии программы. Бэкапы старше 30 дней удаляются.",
            font=(APP_FONT, 14),
            justify="left"
        )
        text.pack(anchor="w", padx=22, pady=(0, 12))
        self.muted_labels.append(text)
        self.settings_buttons_frame = ctk.CTkFrame(settings_card, fg_color="transparent")
        self.settings_buttons_frame.pack(anchor="w", padx=16, pady=2)
        self.transparent_frames.append(self.settings_buttons_frame)
        self.create_button(self.settings_buttons_frame, "Папка с данными", self.open_data_folder, kind="gray", width=170)
        self.create_button(self.settings_buttons_frame, "Очистить историю", self.clear_history, kind="danger", width=180)

        lists_card = self.create_card(self.settings_tab, height=170)
        title = ctk.CTkLabel(lists_card, text="Списки для заполнения", font=(APP_FONT, 22, "bold"))
        title.pack(anchor="w", padx=22, pady=(18, 6))
        self.text_labels.append(title)
        form = ctk.CTkFrame(lists_card, fg_color="transparent")
        form.pack(anchor="w", padx=12, pady=4)
        self.transparent_frames.append(form)
        self.create_label(form, "Что удалить", 0, 0)
        self.list_delete_type_box = self.create_combo(form, ["Этаж", "Расположение", "Принтер", "Расходник"], 1, 0, width=180, command=self.update_delete_value_box, state="readonly")
        self.list_delete_type_box.set("Расположение")
        self.create_label(form, "Значение", 0, 1)
        self.list_delete_value_box = self.create_combo(form, [], 1, 1, width=260, state="readonly")
        delete_frame = ctk.CTkFrame(form, fg_color="transparent")
        delete_frame.grid(row=1, column=2, padx=10, pady=(0, 10), sticky="w")
        self.transparent_frames.append(delete_frame)
        self.create_button(delete_frame, "Удалить из списка", self.delete_selected_list_value, kind="danger", width=180)

        catalog_card = self.create_card(self.settings_tab, fill="both", expand=True)
        title = ctk.CTkLabel(catalog_card, text="Справочник артикулов", font=(APP_FONT, 22, "bold"))
        title.pack(anchor="w", padx=22, pady=(18, 6))
        self.text_labels.append(title)
        text = ctk.CTkLabel(catalog_card, text="Схема: принтер → тип расходника → цвет → артикул.", font=(APP_FONT, 14))
        text.pack(anchor="w", padx=22, pady=(0, 12))
        self.muted_labels.append(text)
        self.catalog_form = ctk.CTkFrame(catalog_card, fg_color="transparent")
        self.catalog_form.pack(anchor="w", padx=12, pady=4)
        self.transparent_frames.append(self.catalog_form)
        self.create_label(self.catalog_form, "Принтер", 0, 0)
        self.catalog_printer_box = self.create_combo(self.catalog_form, self.printer_values(), 1, 0, width=190, command=self.load_catalog_article_to_form)
        self.create_label(self.catalog_form, "Тип", 0, 1)
        self.catalog_type_box = self.create_combo(self.catalog_form, self.type_values(), 1, 1, width=190, command=self.catalog_type_changed)
        self.catalog_type_box.set("Принт-картридж")
        self.create_label(self.catalog_form, "Цвет", 0, 2)
        self.catalog_color_box = self.create_combo(self.catalog_form, COLORS, 1, 2, width=160, command=self.load_catalog_article_to_form, state="readonly")
        self.catalog_color_box.set("Черный")
        self.create_label(self.catalog_form, "Артикул", 0, 3)
        self.catalog_article_entry = self.create_entry(self.catalog_form, 1, 3, width=220)
        self.catalog_save_frame = ctk.CTkFrame(self.catalog_form, fg_color="transparent")
        self.catalog_save_frame.grid(row=1, column=4, padx=10, pady=(0, 10), sticky="w")
        self.transparent_frames.append(self.catalog_save_frame)
        self.create_button(self.catalog_save_frame, "Сохранить артикул", self.save_catalog_article, width=180)
        self.create_button(self.catalog_save_frame, "Удалить артикул", self.delete_catalog_article, kind="danger", width=160)
        self.catalog_preview = ctk.CTkLabel(catalog_card, text="", font=(APP_FONT, 13), justify="left", anchor="w")
        self.catalog_preview.pack(anchor="w", padx=22, pady=(12, 16))
        self.muted_labels.append(self.catalog_preview)
        self.update_delete_value_box()

    # ---------- списки ----------

    def warehouse_values(self):
        return sorted_values(self.lists.get("warehouses", DEFAULT_WAREHOUSES))

    def printer_values(self):
        return sorted_values(self.lists.get("printers", []))

    def type_values(self):
        return sorted_values(self.lists.get("item_types", DEFAULT_ITEM_TYPES))

    def location_values(self, warehouse=None):
        warehouse = warehouse or self.warehouse_box.get().strip()
        return sorted_values(self.lists.get("locations", {}).get(warehouse, []))

    def update_location_values(self, clear_current=False):
        values = self.location_values()
        self.location_box.configure(values=values)
        if clear_current:
            self.location_box.set("")

    def update_all_combo_values(self):
        warehouses = self.warehouse_values()
        printers = self.printer_values()
        types = self.type_values()
        self.warehouse_box.configure(values=warehouses)
        self.printer_box.configure(values=printers)
        self.type_box.configure(values=types)
        self.stock_type_filter.configure(values=["Все"] + types)
        self.catalog_printer_box.configure(values=printers)
        self.catalog_type_box.configure(values=types)
        self.update_location_values(clear_current=False)
        self.rebuild_warehouse_buttons()
        if hasattr(self, "list_delete_value_box"):
            self.update_delete_value_box()

    def rebuild_warehouse_buttons(self):
        for widget in self.warehouse_switch_frame.winfo_children():
            widget.destroy()
        self.warehouse_buttons.clear()
        for name in ["Все"] + self.warehouse_values():
            button = ctk.CTkButton(
                self.warehouse_switch_frame,
                text=name,
                command=lambda value=name: self.set_warehouse_filter(value),
                width=95,
                height=38,
                corner_radius=12,
                font=(APP_FONT, 13, "bold")
            )
            button.pack(side="left", padx=4)
            self.warehouse_buttons[name] = button
        self.update_warehouse_buttons()

    def all_locations(self):
        values = []
        for location_list in self.lists.get("locations", {}).values():
            for value in location_list:
                add_unique(values, value)
        return sorted_values(values)

    def update_delete_value_box(self, value=None):
        if not hasattr(self, "list_delete_value_box"):
            return
        selected = self.list_delete_type_box.get()
        if selected == "Этаж":
            values = self.warehouse_values()
        elif selected == "Расположение":
            values = self.all_locations()
        elif selected == "Принтер":
            values = self.printer_values()
        elif selected == "Расходник":
            values = self.type_values()
        else:
            values = []
        self.list_delete_value_box.configure(values=values)
        self.list_delete_value_box.set(values[0] if values else "")

    def delete_selected_list_value(self):
        list_type = self.list_delete_type_box.get()
        value = self.list_delete_value_box.get().strip()
        if not value:
            messagebox.showwarning("Ошибка", "Выберите значение для удаления.")
            return
        if list_type == "Расходник" and value in DEFAULT_ITEM_TYPES:
            messagebox.showwarning("Ошибка", "Стандартный тип расходника лучше не удалять.")
            return
        if not messagebox.askyesno("Удаление из списка", f"Удалить «{value}» из списка «{list_type}»?\n\nУже созданные позиции склада не удалятся. Значение не вернется автоматически при следующем запуске."):
            return

        if list_type == "Этаж":
            self.mark_deleted_value("warehouses", value)
            if value in self.lists.get("warehouses", []):
                self.lists["warehouses"].remove(value)
            self.lists.get("locations", {}).pop(value, None)
            if self.warehouse_filter == value:
                self.warehouse_filter = "Все"
        elif list_type == "Расположение":
            self.mark_deleted_value("locations", value)
            for locations in self.lists.get("locations", {}).values():
                while value in locations:
                    locations.remove(value)
        elif list_type == "Принтер":
            self.mark_deleted_value("printers", value)
            while value in self.lists.get("printers", []):
                self.lists["printers"].remove(value)
        elif list_type == "Расходник":
            self.mark_deleted_value("item_types", value)
            while value in self.lists.get("item_types", []):
                self.lists["item_types"].remove(value)

        self.sort_lists()
        self.save_lists()
        self.update_all_combo_values()
        messagebox.showinfo("Список", f"Значение «{value}» удалено из списка.")

    # ---------- обновление таблиц ----------

    def update_tree_headings(self):
        for column, title, _ in self.stock_columns:
            arrow = ""
            if column == self.stock_sort_column:
                arrow = " ↓" if self.stock_sort_reverse else " ↑"
            self.stock_tree.heading(column, text=title + arrow, anchor="center", command=lambda c=column: self.sort_stock(c))
        for column, title, _ in self.history_columns:
            arrow = ""
            if column == self.history_sort_column:
                arrow = " ↓" if self.history_sort_reverse else " ↑"
            self.history_tree.heading(column, text=title + arrow, anchor="center", command=lambda c=column: self.sort_history(c))

    def refresh_stock_table(self, keep_selection=False):
        selected_before = self.selected_id if keep_selection else None
        self.stock_tree.delete(*self.stock_tree.get_children())
        for index, item in enumerate(self.get_stock_records()):
            status = get_status(item.get("quantity", 0), item.get("min_quantity", 1))
            tag = "selected_row" if item.get("id") == selected_before else self.get_stock_row_tag(item, index)
            self.stock_tree.insert("", "end", iid=item["id"], values=[
                item.get("warehouse", ""),
                item.get("location", ""),
                item.get("type", ""),
                item.get("color", "-"),
                item.get("article", ""),
                item.get("printer", ""),
                item.get("quantity", 0),
                item.get("min_quantity", 1),
                status
            ], tags=(tag,))
        if selected_before and self.stock_tree.exists(selected_before):
            self.stock_tree.selection_set(selected_before)
            self.stock_tree.focus(selected_before)
            self.refresh_stock_selection_style()
        self.update_tree_headings()

    def refresh_history_table(self):
        self.history_tree.delete(*self.history_tree.get_children())
        for index, record in enumerate(self.get_history_records()):
            tag = "even" if index % 2 == 0 else "odd"
            self.history_tree.insert("", "end", iid=record.get("id", str(uuid4())), values=[
                record.get("date", ""),
                record.get("action", ""),
                record.get("warehouse", ""),
                record.get("location", ""),
                record.get("type", ""),
                record.get("color", "-"),
                record.get("article", ""),
                record.get("printer", ""),
                record.get("before", ""),
                record.get("after", ""),
                record.get("change", ""),
                record.get("comment", "")
            ], tags=(tag,))
        self.update_tree_headings()

    def refresh_stats(self):
        records = self.get_stock_records()
        low = 0
        empty = 0
        for item in records:
            status = get_status(item.get("quantity", 0), item.get("min_quantity", 1))
            if status == "Мало":
                low += 1
            elif status == "Нет":
                empty += 1
        self.stat_labels["total"].configure(text=str(len(records)))
        self.stat_labels["low"].configure(text=str(low))
        self.stat_labels["empty"].configure(text=str(empty))
        self.stat_labels["floor"].configure(text=self.warehouse_filter)

    def update_reports_info(self):
        if not hasattr(self, "reports_info_label"):
            return
        reports_count = len(list(REPORTS_DIR.glob("*.xlsx"))) if REPORTS_DIR.exists() else 0
        backups_count = len(list(BACKUPS_DIR.glob("*.json"))) if BACKUPS_DIR.exists() else 0
        self.reports_info_label.configure(text=f"Отчетов склада: {reports_count} | Файлов бэкапа: {backups_count}")

    def refresh_all(self):
        self.refresh_stats()
        self.refresh_stock_table(keep_selection=True)
        self.refresh_history_table()
        self.update_reports_info()

    # ---------- тема ----------


    def update_warehouse_buttons(self):
        if not hasattr(self, "warehouse_buttons"):
            return
        p = self.theme()
        for name, button in self.warehouse_buttons.items():
            if name == self.warehouse_filter:
                button.configure(fg_color=p["accent"], hover_color=p["accent_hover"], text_color="#FFFFFF")
            else:
                button.configure(fg_color=p["button_gray"], hover_color=p["button_gray_hover"], text_color=p["button_gray_text"])

    def apply_tree_theme(self):
        p = self.theme()
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure(
            "Clean.Treeview",
            background=p["card"],
            fieldbackground=p["card"],
            foreground=p["text"],
            rowheight=42,
            borderwidth=0,
            relief="flat",
            font=(APP_FONT, 11),
            bordercolor=p["card"],
            lightcolor=p["card"],
            darkcolor=p["card"]
        )
        style.configure(
            "Clean.Treeview.Heading",
            background=p["header"],
            foreground=p["muted"],
            borderwidth=0,
            relief="flat",
            font=(APP_FONT, 10, "bold"),
            bordercolor=p["header"],
            lightcolor=p["header"],
            darkcolor=p["header"]
        )
        style.map(
            "Clean.Treeview",
            background=[("selected", p["selected"])],
            foreground=[("selected", p["selected_text"])]
        )
        style.map("Clean.Treeview.Heading", background=[("active", p["selected"])])
        for tree in [self.stock_tree, self.history_tree]:
            tree.configure(style="Clean.Treeview")
            tree.tag_configure("even", background=p["row"], foreground=p["text"])
            tree.tag_configure("odd", background=p["row_alt"], foreground=p["text"])
            tree.tag_configure("ok", background=p["ok"], foreground=p["text"])
            tree.tag_configure("low", background=p["low"], foreground=p["text"])
            tree.tag_configure("empty", background=p["empty"], foreground=p["text"])
            tree.tag_configure("selected_row", background=p["selected"], foreground=p["selected_text"])
        for scrollbar in [self.stock_y_scroll, self.stock_x_scroll, self.history_y_scroll, self.history_x_scroll]:
            scrollbar.configure(fg_color="transparent", button_color=p["scrollbar"], button_hover_color=p["scrollbar_hover"])

    def apply_theme(self):
        p = self.theme()
        ctk.set_appearance_mode("light")
        self.configure(fg_color=p["bg"])
        safe_configure(
            self.tab_view,
            fg_color=p["bg"],
            segmented_button_fg_color=p["header"],
            segmented_button_selected_color=p["accent"],
            segmented_button_selected_hover_color=p["accent_hover"],
            segmented_button_unselected_color=p["header"],
            segmented_button_unselected_hover_color=p["selected"],
            text_color=p["text"]
        )
        for tab in [self.stock_tab, self.history_tab, self.reports_tab, self.settings_tab]:
            safe_configure(tab, fg_color=p["bg"])
        for card in self.cards:
            safe_configure(card, fg_color=p["card"], border_color=p["border"])
        for label in self.text_labels:
            label.configure(text_color=p["text"])
        for label in self.muted_labels:
            label.configure(text_color=p["muted"])
        for widget in self.inputs:
            if isinstance(widget, ctk.CTkComboBox):
                safe_configure(
                    widget,
                    fg_color=p["input"],
                    border_color=p["border"],
                    text_color=p["text"],
                    button_color=p["button_gray"],
                    button_hover_color=p["button_gray_hover"],
                    dropdown_fg_color=p["card"],
                    dropdown_hover_color=p["header"],
                    dropdown_text_color=p["text"]
                )
            else:
                safe_configure(widget, fg_color=p["input"], border_color=p["border"], text_color=p["text"])
        for button, kind in self.buttons:
            if kind == "danger":
                button.configure(fg_color=p["danger"], hover_color=p["danger_hover"], text_color="#FFFFFF")
            elif kind == "success":
                button.configure(fg_color=p["success"], hover_color=p["success_hover"], text_color="#FFFFFF")
            elif kind == "gray":
                button.configure(fg_color=p["button_gray"], hover_color=p["button_gray_hover"], text_color=p["button_gray_text"])
            else:
                button.configure(fg_color=p["accent"], hover_color=p["accent_hover"], text_color="#FFFFFF")
        for frame in self.transparent_frames:
            safe_configure(frame, fg_color="transparent")
        self.update_warehouse_buttons()
        self.apply_tree_theme()
        self.refresh_all()
        self.update_catalog_preview()


if __name__ == "__main__":
    app = App()
    app.mainloop()
